import openpyxl
import json
import os
import glob
import datetime

CURRENT_DIR = os.path.dirname(__file__)
BASE_DIR = os.path.dirname(CURRENT_DIR)
DATE_INDEX = 22

# create the folders
os.makedirs(os.path.join(CURRENT_DIR, 'cache'), exist_ok=True)
os.makedirs(os.path.join(CURRENT_DIR, 'dialer files'), exist_ok=True)

# get the current datetime and create a formatted string
now = datetime.datetime.now()
formatted_time = datetime.datetime.strftime(now, '%Y-%m-%d %H-%M')

xlsx_files_full_path = glob.glob(os.path.join(BASE_DIR, '*.xlsx'))
xlsx_files = [file.split(os.path.sep)[-1] for file in xlsx_files_full_path]

print('select the file you want to import into dialer:')
for i, file in enumerate(xlsx_files):
    print(f'{i + 1}. {file}')

# get the filename to be imported
try:
    file_index = int(input('file number: '))
    selected_file_path = xlsx_files_full_path[file_index - 1]
    selected_file_name = xlsx_files[file_index - 1]
    selected_location_name = selected_file_name[:-5]
except:
    print('enter a valid number')
    quit()

print('creating the new lists...')

# read the file data
wb = openpyxl.load_workbook(filename=selected_file_path)
sheet = wb.active

# get the cache
DIALER_CACHE_FILE = os.path.join(CURRENT_DIR, 'cache', 'dialer_cache.json')

if os.path.exists(DIALER_CACHE_FILE):
    with open(DIALER_CACHE_FILE, encoding='utf-8') as f:
        dialer_cache = json.load(f)
else:
    dialer_cache = {}

# if the cache schema doesn't exist, then create it
if not dialer_cache.get(selected_location_name):
    dialer_cache[selected_location_name] = {
        'new': [], 'sold': []
    }

existing_new_properties = dialer_cache[selected_location_name]['new']
existing_sold_properties = dialer_cache[selected_location_name]['sold']

new_properties = []
sold_properties = []
headers = None

for i, row in enumerate(sheet.iter_rows()):
    if i == 0:
        headers = [cell.value for cell in row][:-1]
        continue

    _id = row[0].value
    name = row[9].value
    identifier = f'{_id}-{name}'.replace(' ', '_')
    row = [cell.value for cell in row]
    date = row[DATE_INDEX]

    # check for new property
    if identifier not in existing_new_properties:
        dialer_cache[selected_location_name]['new'].append(identifier)
        new_properties.append(row[:-1])

    if _id not in existing_sold_properties and date:
        dialer_cache[selected_location_name]['sold'].append(_id)
        sold_properties.append([_id, date])

# save the cache
with open(DIALER_CACHE_FILE, 'w', encoding='utf-8') as f:
    json.dump(dialer_cache, f, indent=2)

# create the file names and locations for the files to be saved
new_file_name = f'{selected_location_name} [{formatted_time}] (new).xlsx'
sold_file_name = f'{selected_location_name} [{formatted_time}] (sold).xlsx'
new_file_path = os.path.join(CURRENT_DIR, 'dialer files', new_file_name)
sold_file_path = os.path.join(CURRENT_DIR, 'dialer files', sold_file_name)

# save the new properties
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.append(headers)
for row in new_properties:
    new_sheet.append(row)
new_wb.save(new_file_path)

# save the sold properties
sold_wb = openpyxl.Workbook()
sold_sheet = sold_wb.active
sold_sheet.append(['Id', 'Date'])
for row in sold_properties:
    sold_sheet.append(row)
sold_wb.save(sold_file_path)

print('complete!')